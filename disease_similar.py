# 在disease_vector_max/average存在的情况下  计算positivate_example和random_example的相似度

import numpy as np
from openpyxl import load_workbook
import pickle


def load_obj(name):
    with open(name+'.pkl', 'rb') as f:
        return pickle.load(f)


# 计算向量的余弦相似度
def similar(a, b):+.0


    N = [[None]]
    vector_a = np.mat(a)
    vector_b = np.mat(b)
    try:
        num1 = float(vector_a * vector_b.T)  # 矩阵相乘
        denom1 = np.linalg.norm(vector_a) * np.linalg.norm(vector_b)
    except:
        return 0

    if denom1 != 0:
        cos1 = num1 / denom1
        return cos1
    else:
        return -1

fpkl = load_obj('data/disease_embedding')


#获得相似度
def get_data(path):
    # 对正例计算相似度
    workbook = load_workbook(path)  # 找到需要xlsx文件的位置
    booksheet = workbook.active  # 获取当前活跃的sheet,默认是第一个sheet
    f = open('data/random2_similar.txt', 'a+', encoding='utf-8')
    # 获取sheet页的行数据
    rows = booksheet.rows
    i = 0
    # 迭代所有的行
    for _ in rows:
        i = i + 1
        cell_data_1 = booksheet.cell(row=i, column=2).value  # 获取第i行1 列的数据
        cell_data_2 = booksheet.cell(row=i, column=4).value  # 获取第i行 2 列的数据
        #获得疾病名对应的疾病向量6
        try:
            # 获得疾病名对应的疾病向量6
            vector1 = fpkl[cell_data_1]
        except:
            print(cell_data_1, "not exit")
            continue
        try:
            # 获得疾病名对应的疾病向量6
            vector2 = fpkl[cell_data_2]
        except:
            print(cell_data_2, "not exit")
            continue
        s = similar(vector1, vector2)
        if s != -1:
            f.write(cell_data_1)
            f.write(',')
            f.write(cell_data_2)
            f.write(':')
            f.write(str(s))
            f.write('\n')
    f.close()
    workbook.close()

    return 0

get_data("data/random2.xlsx")
